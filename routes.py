import csv
import io
import os
import requests
import time
from urllib.parse import quote, urlparse
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from flask import render_template, request, redirect, url_for, flash, make_response, jsonify, session
from functools import wraps
from flask_login import current_user
from app import app, db

# Импорты
from models import Flower, Category, Bouquet, BouquetComposition, GlobalSetting, User
from forms import FlowerForm, CategoryForm, BouquetForm, GlobalSettingsForm, ExportForm, LoginForm, UserForm, RegisterForm


def build_bouquet_filter_query(category_id=None, search_query='', published_filter=''):
    query = Bouquet.query

    if category_id:
        query = query.filter_by(category_id=category_id)

    if published_filter == '1':
        query = query.filter_by(published=True)
    elif published_filter == '0':
        query = query.filter_by(published=False)

    if search_query:
        search_pattern = f'%{search_query}%'
        query = query.filter(
            (Bouquet.name.ilike(search_pattern)) |
            (Bouquet.sku.ilike(search_pattern))
        )

    return query


def get_site_sync_settings():
    endpoint = GlobalSetting.get_value(
        'site_sync_endpoint',
        None,
        os.environ.get('SITE_SYNC_ENDPOINT', '')
    )
    token = GlobalSetting.get_value(
        'site_sync_token',
        None,
        os.environ.get('SITE_SYNC_TOKEN', '')
    )
    host_header = GlobalSetting.get_value(
        'site_sync_host_header',
        None,
        os.environ.get('SITE_SYNC_HOST_HEADER', '')
    )
    verify_ssl = GlobalSetting.get_value(
        'site_sync_verify_ssl',
        None,
        os.environ.get('SITE_SYNC_VERIFY_SSL', '1')
    )

    return {
        'endpoint': (endpoint or '').strip(),
        'token': (token or '').strip(),
        'host_header': (host_header or '').strip(),
        'verify_ssl': str(verify_ssl).lower() not in ('0', 'false', 'no', 'off'),
    }


def sync_prices_to_site(items):
    settings = get_site_sync_settings()
    if not settings['endpoint'] or not settings['token']:
        return {
            'success': False,
            'configured': False,
            'message': 'Синхронизация с сайтом еще не настроена: нужен endpoint Joomla и секретный token.',
            'results': []
        }

    headers = {
        'X-CRM-Token': settings['token'],
        'Content-Type': 'application/json'
    }

    if settings['host_header']:
        headers['Host'] = settings['host_header']

    response = requests.post(
        settings['endpoint'],
        json={'items': items},
        headers=headers,
        verify=settings['verify_ssl'],
        timeout=30
    )

    try:
        data = response.json()
    except ValueError:
        data = {'message': response.text[:500]}

    return {
        'success': response.ok and bool(data.get('success', True)),
        'configured': True,
        'status_code': response.status_code,
        'message': data.get('message', 'Ответ сайта получен.'),
        'results': data.get('results', [])
    }

# Проверка авторизации
def verify_live_site_prices(items, sync_result):
    """Check that the public storefront renders the new prices after endpoint update."""
    if not sync_result.get('success') or not sync_result.get('results'):
        return sync_result

    settings = get_site_sync_settings()
    public_host = settings['host_header'] or urlparse(settings['endpoint']).hostname
    if not public_host:
        return sync_result

    item_by_sku = {item['sku']: item for item in items}
    verified = []
    mismatched = []
    skipped = 0

    headers = {
        'Cache-Control': 'no-cache, no-store, max-age=0',
        'Pragma': 'no-cache',
        'User-Agent': 'FlowersMafiaCRM-LivePriceCheck/1.0',
    }

    for result in sync_result.get('results', [])[:10]:
        if not result.get('success'):
            continue

        sku = result.get('sku')
        item = item_by_sku.get(sku)
        product_id = result.get('product_id')

        if not item or not product_id:
            skipped += 1
            continue

        expected_price = int(item['price'])
        check_url = (
            f'https://{public_host}/index.php?option=com_virtuemart'
            f'&view=productdetails&virtuemart_product_id={int(product_id)}'
            f'&crmcheck={int(time.time())}'
        )

        try:
            response = requests.get(check_url, headers=headers, timeout=20)
            page = response.text
            expected_tokens = [
                f'"price":"{expected_price}.00"',
                f'"price": "{expected_price}.00"',
                f"content='{expected_price}'",
                f'content="{expected_price}"',
                f'>{expected_price} MDL',
            ]

            if response.ok and any(token in page for token in expected_tokens):
                verified.append({
                    'sku': sku,
                    'price': expected_price,
                    'url': response.url,
                })
            else:
                mismatched.append({
                    'sku': sku,
                    'price': expected_price,
                    'url': response.url,
                    'status_code': response.status_code,
                    'message': 'Живой сайт не показывает новую цену.',
                })
        except Exception as exc:
            mismatched.append({
                'sku': sku,
                'price': expected_price,
                'url': check_url,
                'message': str(exc),
            })

    sync_result['live_check'] = {
        'checked': len(verified) + len(mismatched),
        'verified': verified,
        'mismatched': mismatched,
        'skipped': skipped,
    }

    if mismatched:
        sync_result['success'] = False
        sync_result['message'] = (
            'Endpoint принял цены, но живой сайт не подтвердил обновление. '
            'Публичный домен смотрит на другой сервер или кэш.'
        )

    return sync_result


def require_login(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user_id' not in session:
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated_function

# Глобальная проверка авторизации для всех маршрутов
OPEN_ROUTES = {'login', 'logout', 'register', 'api_login', 'static'}

@app.before_request
def check_auth():
    if request.endpoint and request.endpoint not in OPEN_ROUTES:
        if 'user_id' not in session:
            return redirect(url_for('login'))

# Authentication routes
@app.route('/login', methods=['GET', 'POST'])
def login():
    """Login page"""
    if 'user_id' in session:
        return redirect(url_for('index'))
    
    form = LoginForm()
    if form.validate_on_submit():
        try:
            user = User.query.filter_by(username=form.username.data).first()
            print(f"Попытка входа: логин='{form.username.data}', найден пользователь: {user is not None}")
            if user:
                password_check = user.check_password(form.password.data)
                print(f"Проверка пароля: {password_check}")
            
            if user and user.check_password(form.password.data):
                print(f"Успешная аутентификация для пользователя: {user.username}")
                
                # Проверяем статус пользователя
                if user.status != 'active':
                    flash('Ваш аккаунт ожидает одобрения администратора', 'warning')
                    return render_template('login.html', form=form)
                
                # Сохраняем пользователя в сессии
                session['user_id'] = user.id
                session['username'] = user.username
                session['role'] = user.role
                session.permanent = True
                
                print(f"Сохранили в сессию: user_id={user.id}, username={user.username}")
                
                flash('Добро пожаловать в систему!', 'success')
                return redirect(url_for('index'))
            else:
                flash('Неверный логин или пароль', 'error')
        except Exception as e:
            print(f"Ошибка при входе: {e}")
            db.session.rollback()
            flash('Произошла ошибка при входе в систему', 'error')
        finally:
            db.session.close()
    
    return render_template('login.html', form=form)

@app.route('/logout')
def logout():
    """Logout user"""
    # Очищаем сессию полностью
    session.clear()
    flash('Вы успешно вышли из системы', 'info')
    return redirect(url_for('login'))

@app.route('/register')
def register():
    """Registration disabled"""
    return redirect(url_for('login'))

@app.route('/')
def index():
    """Главная страница с полной админкой"""
    # Получаем данные для дашборда
    flower_count = Flower.query.count()
    category_count = Category.query.count()
    bouquet_count = Bouquet.query.filter_by(published=True).count()
    recent_bouquets = Bouquet.query.filter_by(published=True).order_by(Bouquet.id.desc()).limit(10).all()
    
    return render_template('index.html', 
                         flower_count=flower_count,
                         category_count=category_count,
                         bouquet_count=bouquet_count,
                         recent_bouquets=recent_bouquets,
                         user_logged_in=True)

# API для входа в систему
@app.route('/api/login', methods=['POST'])
def api_login():
    """API endpoint для входа"""
    try:
        data = request.get_json()
        username = data.get('username')
        password = data.get('password')
        
        print(f"API вход: логин='{username}'")
        
        user = User.query.filter_by(username=username).first()
        
        if user and user.check_password(password):
            print(f"API: Успешная аутентификация для {username}")
            return jsonify({'success': True, 'message': 'Добро пожаловать!'})
        else:
            print(f"API: Неверные данные для {username}")
            return jsonify({'success': False, 'message': 'Неверный логин или пароль'})
    
    except Exception as e:
        print(f"Ошибка API входа: {e}")
        return jsonify({'success': False, 'message': 'Ошибка сервера'})

# API для получения данных дашборда
@app.route('/api/dashboard')
def api_dashboard():
    """API endpoint для дашборда"""
    try:
        flower_count = Flower.query.count()
        category_count = Category.query.count()
        bouquet_count = Bouquet.query.count()
        
        return jsonify({
            'flowers_count': flower_count,
            'categories_count': category_count,
            'bouquets_count': bouquet_count
        })
    except Exception as e:
        print(f"Ошибка API дашборда: {e}")
        return jsonify({'error': 'Ошибка загрузки данных'})

# API для получения настроек
@app.route('/api/settings')
def api_settings():
    """API endpoint для настроек"""
    try:
        # Временно используем админа (ID=1) для настроек
        admin_user = User.query.filter_by(username='admin').first()
        if not admin_user:
            return jsonify({
                'delivery_cost': 100.0,
                'markup_percentage': 50.0
            })
            
        delivery_cost_value = GlobalSetting.get_value('delivery_cost', admin_user.id, '100.0')
        markup_percentage_value = GlobalSetting.get_value('markup_percentage', admin_user.id, '50.0')
        
        delivery_cost = float(delivery_cost_value) if delivery_cost_value else 100.0
        markup_percentage = float(markup_percentage_value) if markup_percentage_value else 50.0
        
        return jsonify({
            'delivery_cost': delivery_cost,
            'markup_percentage': markup_percentage
        })
    except Exception as e:
        print(f"Ошибка API настроек: {e}")
        return jsonify({
            'delivery_cost': 100.0,
            'markup_percentage': 50.0
        })

# API для управления цветами
@app.route('/api/flowers')
def api_flowers():
    """Получить список всех цветов"""
    try:
        flowers = Flower.query.order_by(Flower.name).all()
        return jsonify([{
            'id': f.id,
            'name': f.name,
            'price_per_unit': f.price_per_unit
        } for f in flowers])
    except Exception as e:
        print(f"Ошибка API цветов: {e}")
        return jsonify({'error': 'Ошибка загрузки цветов'})

@app.route('/api/flowers', methods=['POST'])
def api_flowers_create():
    """Создать новый цветок"""
    try:
        data = request.get_json()
        flower = Flower(
            name=data['name'],
            price_per_unit=float(data['price_per_unit']),
            user_id=session.get('user_id', 3)
        )
        db.session.add(flower)
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': f'Цветок "{flower.name}" добавлен',
            'id': flower.id
        })
    except Exception as e:
        print(f"Ошибка создания цветка: {e}")
        return jsonify({'success': False, 'message': 'Ошибка создания цветка'})

@app.route('/api/flowers/<int:flower_id>', methods=['PUT'])
def api_flowers_update(flower_id):
    """Обновить цветок"""
    try:
        flower = Flower.query.get_or_404(flower_id)
        data = request.get_json()
        
        flower.name = data['name']
        flower.price_per_unit = float(data['price_per_unit'])
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': f'Цветок "{flower.name}" обновлен'
        })
    except Exception as e:
        print(f"Ошибка обновления цветка: {e}")
        return jsonify({'success': False, 'message': 'Ошибка обновления цветка'})

@app.route('/api/flowers/<int:flower_id>', methods=['DELETE'])
def api_flowers_delete(flower_id):
    """Удалить цветок"""
    try:
        flower = Flower.query.get_or_404(flower_id)
        name = flower.name
        db.session.delete(flower)
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': f'Цветок "{name}" удален'
        })
    except Exception as e:
        print(f"Ошибка удаления цветка: {e}")
        return jsonify({'success': False, 'message': 'Ошибка удаления цветка'})

# Простая проверка через cookies
def check_auth():
    auth_token = request.cookies.get('auth_token')
    print(f"Проверка cookie: auth_token={auth_token}")
    if not auth_token or not auth_token.startswith('user_'):
        print("Cookie отсутствует или неверный, перенаправляем на логин")
        return redirect(url_for('login'))
    print("Cookie действителен, доступ разрешен")
    return None

# Flower routes
@app.route('/flowers')
def flowers_index():
    """List all flowers"""
    print(f"=== ЦВЕТЫ === Загружаем страницу")
    flowers = Flower.query.order_by(Flower.name).all()
    return render_template('flowers/index.html', flowers=flowers)

@app.route('/flowers/new', methods=['GET', 'POST'])
def flowers_new():
    """Create a new flower"""
    form = FlowerForm()
    
    if request.method == 'POST':
        action = request.form.get('action', 'save')
        
        if form.validate_on_submit():
            # Check if flower name already exists
            existing_flower = Flower.query.filter_by(name=form.name.data).first()
            if existing_flower:
                flash('Цветок с таким названием уже существует.', 'error')
                return render_template('flowers/form.html', form=form, title='Добавить новый цветок')
            
            flower = Flower(
                name=form.name.data,
                price_per_unit=form.price_per_unit.data,
                user_id=session.get('user_id', 3)
            )
            
            try:
                db.session.add(flower)
                db.session.commit()
                flash(f'Цветок "{flower.name}" успешно добавлен!', 'success')
                
                if action == 'save_and_close':
                    return redirect(url_for('flowers_index'))
                else:
                    # Stay on the same page for adding another flower
                    form = FlowerForm()  # Reset form
                    
            except Exception as e:
                db.session.rollback()
                flash('Ошибка при добавлении цветка. Попробуйте снова.', 'error')
                app.logger.error(f'Error adding flower: {e}')
        else:
            flash('Пожалуйста, исправьте ошибки в форме.', 'error')
    
    return render_template('flowers/form.html', form=form, title='Добавить новый цветок')

@app.route('/flowers/<int:id>/edit', methods=['GET', 'POST'])
def flowers_edit(id):
    """Edit an existing flower"""
    flower = Flower.query.get_or_404(id)
    form = FlowerForm(obj=flower)
    
    if request.method == 'POST':
        action = request.form.get('action', 'save')
        
        if form.validate_on_submit():
            # Check if flower name already exists (excluding current flower)
            existing_flower = Flower.query.filter(
                Flower.name == form.name.data,
                Flower.id != id
            ).first()
            if existing_flower:
                flash('Цветок с таким названием уже существует.', 'error')
                return render_template('flowers/form.html', form=form, title=f'Редактировать {flower.name}')
            
            # Remember old price to check if it changed
            old_price = flower.price_per_unit
            
            flower.name = form.name.data
            flower.price_per_unit = form.price_per_unit.data
            
            try:
                db.session.commit()
                
                # Check if price changed and show affected bouquets count
                if old_price != flower.price_per_unit:
                    # Count affected bouquets
                    affected_bouquets = db.session.query(Bouquet).join(BouquetComposition).filter(
                        BouquetComposition.flower_id == flower.id
                    ).count()
                    
                    if affected_bouquets > 0:
                        flash(f'Цветок "{flower.name}" обновлен! Автоматически пересчитано {affected_bouquets} букетов.', 'success')
                    else:
                        flash(f'Цветок "{flower.name}" обновлен!', 'success')
                else:
                    flash(f'Цветок "{flower.name}" успешно обновлен!', 'success')
                
                print(f"Действие: {action}")
                
                if action == 'save_and_stay':
                    print("Остаемся на странице редактирования")
                    # Stay on current page
                else:
                    print("Перенаправляем к списку цветов")
                    return redirect(url_for('flowers_index'))
                    
            except Exception as e:
                db.session.rollback()
                flash('Ошибка при обновлении цветка. Попробуйте снова.', 'error')
                app.logger.error(f'Error updating flower: {e}')
        else:
            flash('Пожалуйста, исправьте ошибки в форме.', 'error')
    
    return render_template('flowers/form.html', form=form, title=f'Редактировать {flower.name}')

@app.route('/flowers/<int:id>/delete', methods=['POST'])
def flowers_delete(id):
    """Delete a flower"""
    flower = Flower.query.get_or_404(id)
    
    # Check if flower is used in any bouquets
    if flower.compositions:
        bouquet_names = [comp.bouquet.name for comp in flower.compositions]
        flash(f'Нельзя удалить цветок "{flower.name}". Он используется в букетах: {", ".join(bouquet_names)}', 'error')
        return redirect(url_for('flowers_index'))
    
    try:
        db.session.delete(flower)
        db.session.commit()
        flash(f'Цветок "{flower.name}" успешно удален!', 'success')
    except Exception as e:
        db.session.rollback()
        flash('Ошибка при удалении цветка. Попробуйте снова.', 'error')
        app.logger.error(f'Error deleting flower: {e}')
    
    return redirect(url_for('flowers_index'))

# Category routes
@app.route('/categories')
def categories_index():
    """List all categories"""
    categories = Category.query.order_by(Category.name).all()
    return render_template('categories/index.html', categories=categories)

@app.route('/categories/new', methods=['GET', 'POST'])
def categories_new():
    """Create a new category"""
    form = CategoryForm()
    
    if request.method == 'POST':
        action = request.form.get('action', 'save')
        
        if form.validate_on_submit():
            # Check if category name already exists
            existing_category = Category.query.filter_by(name=form.name.data).first()
            if existing_category:
                flash('Категория с таким названием уже существует.', 'error')
                return render_template('categories/form.html', form=form, title='Добавить новую категорию')
            
            category = Category(name=form.name.data, user_id=session.get('user_id', 3))
            
            try:
                db.session.add(category)
                db.session.commit()
                flash(f'Категория "{category.name}" успешно добавлена!', 'success')
                
                if action == 'save_and_close':
                    return redirect(url_for('categories_index'))
                else:
                    # Stay on the same page for adding another category
                    form = CategoryForm()  # Reset form
                    
            except Exception as e:
                db.session.rollback()
                flash('Ошибка при добавлении категории. Попробуйте снова.', 'error')
                app.logger.error(f'Error adding category: {e}')
        else:
            flash('Пожалуйста, исправьте ошибки в форме.', 'error')
    
    return render_template('categories/form.html', form=form, title='Добавить новую категорию')

@app.route('/categories/<int:id>/edit', methods=['GET', 'POST'])
def categories_edit(id):
    """Edit an existing category"""
    category = Category.query.get_or_404(id)
    form = CategoryForm(obj=category)
    
    if request.method == 'POST':
        action = request.form.get('action', 'save')
        
        if form.validate_on_submit():
            # Check if category name already exists (excluding current category)
            existing_category = Category.query.filter(
                Category.name == form.name.data,
                Category.id != id
            ).first()
            if existing_category:
                flash('Категория с таким названием уже существует.', 'error')
                return render_template('categories/form.html', form=form, title=f'Редактировать {category.name}')
            
            category.name = form.name.data
            
            try:
                db.session.commit()
                flash(f'Категория "{category.name}" успешно обновлена!', 'success')
                
                if action == 'save_and_close':
                    return redirect(url_for('categories_index'))
                # If just save, stay on edit page
            except Exception as e:
                db.session.rollback()
                flash('Ошибка при обновлении категории. Попробуйте снова.', 'error')
                app.logger.error(f'Error updating category: {e}')
        else:
            flash('Пожалуйста, исправьте ошибки в форме.', 'error')
    
    return render_template('categories/form.html', form=form, title=f'Edit {category.name}')

@app.route('/categories/<int:id>/delete', methods=['POST'])
def categories_delete(id):
    """Delete a category"""
    category = Category.query.get_or_404(id)
    
    # Check if category has bouquets
    if category.bouquets:
        bouquet_names = [bouquet.name for bouquet in category.bouquets]
        flash(f'Нельзя удалить категорию "{category.name}". В ней есть букеты: {", ".join(bouquet_names)}', 'error')
        return redirect(url_for('categories_index'))
    
    try:
        db.session.delete(category)
        db.session.commit()
        flash(f'Категория "{category.name}" успешно удалена!', 'success')
    except Exception as e:
        db.session.rollback()
        flash('Ошибка при удалении категории. Попробуйте снова.', 'error')
        app.logger.error(f'Error deleting category: {e}')
    
    return redirect(url_for('categories_index'))

# Bouquet routes
@app.route('/bouquets')
def bouquets_index():
    """List all bouquets with pagination and search"""
    # Получаем параметры из URL
    category_id = request.args.get('category_id', type=int)
    search_query = request.args.get('search', '').strip()
    published_filter = request.args.get('published', '')
    page = request.args.get('page', 1, type=int)
    per_page = 20  # Количество букетов на странице
    
    # Строим запрос с фильтрами
    query = build_bouquet_filter_query(category_id, search_query, published_filter)
    
    # Применяем сортировку и пагинацию
    bouquets_paginated = query.order_by(Bouquet.name).paginate(
        page=page, 
        per_page=per_page, 
        error_out=False
    )
    
    # Получаем все категории для фильтра
    categories = Category.query.order_by(Category.name).all()
    
    # Определяем название выбранной категории
    selected_category_name = None
    if category_id:
        selected_category = Category.query.get(category_id)
        if selected_category:
            selected_category_name = selected_category.name
    
    return render_template('bouquets/index.html', 
                         bouquets=bouquets_paginated.items,
                         pagination=bouquets_paginated,
                         categories=categories,
                         selected_category_id=category_id,
                         selected_category_name=selected_category_name,
                         search_query=search_query,
                         published_filter=published_filter)


@app.route('/bouquets/new', methods=['GET', 'POST'])
def bouquets_new():
    """Create a new bouquet"""
    form = BouquetForm()
    
    if request.method == 'POST':
        action = request.form.get('action', 'save')
        
        if form.validate_on_submit():
            bouquet = Bouquet(
                name=form.name.data,
                sku=form.sku.data,
                category_id=form.category_id.data,
                published=form.published.data,
                user_id=session.get('user_id', 3)
            )
            
            try:
                db.session.add(bouquet)
                db.session.flush()  # Get the bouquet ID
                
                # Get composition data from request
                composition_data = request.form.get('composition_data')
                print(f"Получены данные композиции: {composition_data}")
                
                if composition_data:
                    import json
                    compositions = json.loads(composition_data)
                    print(f"Разобранные композиции: {compositions}")
                    
                    for comp in compositions:
                        if comp.get('flower_id') and comp.get('quantity', 0) > 0:
                            composition = BouquetComposition(
                                bouquet_id=bouquet.id,
                                flower_id=int(comp['flower_id']),
                                quantity=int(comp['quantity'])
                            )
                            db.session.add(composition)
                            print(f"Добавлена композиция: цветок {comp['flower_id']}, количество {comp['quantity']}")
                else:
                    print("Данные композиции не получены!")
                
                # Calculate and set final price
                bouquet.update_price()
                
                db.session.commit()
                flash(f'Букет "{bouquet.name}" успешно добавлен!', 'success')
                
                if action == 'save_and_close':
                    # Сохраняем фильтр по категории при возврате к списку
                    category_filter = request.args.get('category_id')
                    if category_filter:
                        return redirect(url_for('bouquets_index', category_id=category_filter))
                    else:
                        return redirect(url_for('bouquets_index'))
                else:
                    # Reset form for new bouquet
                    form = BouquetForm()
                    
            except Exception as e:
                db.session.rollback()
                flash('Ошибка при добавлении букета. Попробуйте снова.', 'error')
                app.logger.error(f'Error adding bouquet: {e}')
        else:
            flash('Пожалуйста, исправьте ошибки в форме.', 'error')
    
    return render_template('bouquets/form.html', form=form, title='Добавить новый букет')

@app.route('/bouquets/<int:id>/edit', methods=['GET', 'POST'])
def bouquets_edit(id):
    """Edit an existing bouquet"""
    bouquet = Bouquet.query.get_or_404(id)
    form = BouquetForm(obj=bouquet)
    
    if request.method == 'POST':
        action = request.form.get('action', 'save')
        print(f"Действие в редактировании букета: {action}")
        
        if form.validate_on_submit():
            bouquet.name = form.name.data
            bouquet.sku = form.sku.data
            bouquet.category_id = form.category_id.data
            bouquet.published = form.published.data
            
            try:
                # Remove existing compositions
                BouquetComposition.query.filter_by(bouquet_id=bouquet.id).delete()
                
                # Get composition data from request
                composition_data = request.form.get('composition_data')
                print(f"Получены данные композиции при редактировании: {composition_data}")
                
                if composition_data:
                    import json
                    compositions = json.loads(composition_data)
                    print(f"Разобранные композиции при редактировании: {compositions}")
                    
                    for comp in compositions:
                        if comp.get('flower_id') and comp.get('quantity', 0) > 0:
                            composition = BouquetComposition(
                                bouquet_id=bouquet.id,
                                flower_id=int(comp['flower_id']),
                                quantity=int(comp['quantity'])
                            )
                            db.session.add(composition)
                            print(f"Обновлена композиция: цветок {comp['flower_id']}, количество {comp['quantity']}")
                else:
                    print("Данные композиции не получены при редактировании!")
                
                # Update final price
                bouquet.update_price()
                
                db.session.commit()
                flash(f'Букет "{bouquet.name}" успешно обновлен!', 'success')
                
                if action == 'save_and_close':
                    print("Перенаправляем на страницу букетов...")
                    # Сохраняем фильтр по категории и поисковый запрос при возврате к списку
                    category_filter = request.args.get('category_id')
                    search_query = request.args.get('search')
                    page = request.args.get('page', 1)
                    
                    redirect_params = {}
                    if category_filter:
                        redirect_params['category_id'] = category_filter
                    if search_query:
                        redirect_params['search'] = search_query
                    if page and page != '1':
                        redirect_params['page'] = page
                    
                    return redirect(url_for('bouquets_index', **redirect_params))
                else:
                    print("Остаемся на странице редактирования...")
                # If just save, stay on edit page
            except Exception as e:
                db.session.rollback()
                flash('Ошибка при обновлении букета. Попробуйте еще раз.', 'error')
                app.logger.error(f'Error updating bouquet: {e}')
        else:
            flash('Пожалуйста, исправьте ошибки в форме.', 'error')
    
    return render_template('bouquets/form.html', form=form, title='Редактировать букет', bouquet=bouquet)


@app.route('/bouquets/<int:id>/clone')
@require_login
def bouquets_clone(id):
    """Clone an existing bouquet"""
    try:
        original_bouquet = Bouquet.query.get_or_404(id)
        
        # Create new bouquet with copied data
        new_bouquet = Bouquet(
            name=f"Копия: {original_bouquet.name}",
            sku=f"{original_bouquet.sku}-copy",
            category_id=original_bouquet.category_id,
            final_price=original_bouquet.final_price,
            published=False,
            user_id=original_bouquet.user_id or session.get('user_id', 3)
        )
        
        db.session.add(new_bouquet)
        db.session.commit()  # Commit to get the new bouquet ID
        
        # Copy all compositions
        for original_comp in original_bouquet.compositions:
            new_composition = BouquetComposition(
                bouquet_id=new_bouquet.id,
                flower_id=original_comp.flower_id,
                quantity=original_comp.quantity
            )
            db.session.add(new_composition)
        
        db.session.commit()
        flash(f'Букет "{original_bouquet.name}" успешно скопирован! Теперь вы можете его отредактировать.', 'success')
        
        # Redirect to edit the new bouquet with preserved search parameters
        category_filter = request.args.get('category_id')
        search_query = request.args.get('search')
        page = request.args.get('page', 1)
        
        redirect_params = {'id': new_bouquet.id}
        if category_filter:
            redirect_params['category_id'] = category_filter
        if search_query:
            redirect_params['search'] = search_query
        if page and page != '1':
            redirect_params['page'] = page
        
        return redirect(url_for('bouquets_edit', **redirect_params))
        
    except Exception as e:
        db.session.rollback()
        app.logger.error(f'Error cloning bouquet: {e}')
        flash('Ошибка при копировании букета. Попробуйте еще раз.', 'error')
        return redirect(url_for('bouquets_index'))


@app.route('/bouquets/<int:id>/delete', methods=['POST'])
def bouquets_delete(id):
    """Delete a bouquet"""
    bouquet = Bouquet.query.get_or_404(id)
    
    try:
        db.session.delete(bouquet)
        db.session.commit()
        flash(f'Bouquet "{bouquet.name}" deleted successfully!', 'success')
    except Exception as e:
        db.session.rollback()
        flash('Error deleting bouquet. Please try again.', 'error')
        app.logger.error(f'Error deleting bouquet: {e}')
    
    # Сохраняем фильтр по категории при возврате к списку
    category_filter = request.args.get('category_id')
    if category_filter:
        return redirect(url_for('bouquets_index', category_id=category_filter))
    else:
        return redirect(url_for('bouquets_index'))

@app.route('/bouquets/<int:id>/publish', methods=['POST'])
def bouquets_publish(id):
    """Publish a bouquet"""
    bouquet = Bouquet.query.get_or_404(id)
    
    try:
        bouquet.published = True
        db.session.commit()
        flash(f'Букет "{bouquet.name}" опубликован!', 'success')
    except Exception as e:
        db.session.rollback()
        flash('Ошибка при публикации букета. Попробуйте еще раз.', 'error')
        app.logger.error(f'Error publishing bouquet: {e}')
    
    # Сохраняем фильтр по категории при возврате к списку
    category_filter = request.args.get('category_id')
    if category_filter:
        return redirect(url_for('bouquets_index', category_id=category_filter))
    else:
        return redirect(url_for('bouquets_index'))

@app.route('/bouquets/<int:id>/unpublish', methods=['POST'])
def bouquets_unpublish(id):
    """Unpublish a bouquet"""
    bouquet = Bouquet.query.get_or_404(id)
    
    try:
        bouquet.published = False
        db.session.commit()
        flash(f'Букет "{bouquet.name}" снят с публикации!', 'info')
    except Exception as e:
        db.session.rollback()
        flash('Ошибка при снятии с публикации. Попробуйте еще раз.', 'error')
        app.logger.error(f'Error unpublishing bouquet: {e}')
    
    # Сохраняем фильтр по категории при возврате к списку
    category_filter = request.args.get('category_id')
    if category_filter:
        return redirect(url_for('bouquets_index', category_id=category_filter))
    else:
        return redirect(url_for('bouquets_index'))

# Settings routes
@app.route('/settings', methods=['GET', 'POST'])
def settings_index():
    """Global settings page"""
    form = GlobalSettingsForm()
    
    if request.method == 'POST' and form.validate_on_submit():
        try:
            # Проверяем, что данные не None
            delivery_cost = form.delivery_cost.data if form.delivery_cost.data is not None else 600.0
            markup_percentage = form.markup_percentage.data if form.markup_percentage.data is not None else 17.0
            site_sync_endpoint = (form.site_sync_endpoint.data or '').strip()
            site_sync_token = (form.site_sync_token.data or '').strip()
            site_sync_host_header = (form.site_sync_host_header.data or '').strip()
            site_sync_verify_ssl = '1' if form.site_sync_verify_ssl.data else '0'
            
            # Сохраняем глобальные настройки (set_value автоматически сохранит их с user_id=None)
            GlobalSetting.set_value('delivery_cost', str(delivery_cost))
            GlobalSetting.set_value('markup_percentage', str(markup_percentage))
            GlobalSetting.set_value('site_sync_endpoint', site_sync_endpoint)
            GlobalSetting.set_value('site_sync_token', site_sync_token)
            GlobalSetting.set_value('site_sync_host_header', site_sync_host_header)
            GlobalSetting.set_value('site_sync_verify_ssl', site_sync_verify_ssl)
            flash('Настройки успешно обновлены!', 'success')
            return redirect(url_for('settings_index'))
        except Exception as e:
            db.session.rollback()
            flash('Ошибка при обновлении настроек. Попробуйте снова.', 'error')
            app.logger.error(f'Error updating settings: {e}')
    
    # Load current values (get_value автоматически найдет глобальные настройки)
    if request.method == 'GET':
        delivery_cost = GlobalSetting.get_value('delivery_cost', None, '600')
        markup_percentage = GlobalSetting.get_value('markup_percentage', None, '17')
        site_sync_endpoint = GlobalSetting.get_value('site_sync_endpoint', None, os.environ.get('SITE_SYNC_ENDPOINT', ''))
        site_sync_token = GlobalSetting.get_value('site_sync_token', None, os.environ.get('SITE_SYNC_TOKEN', ''))
        site_sync_host_header = GlobalSetting.get_value('site_sync_host_header', None, os.environ.get('SITE_SYNC_HOST_HEADER', ''))
        site_sync_verify_ssl = GlobalSetting.get_value('site_sync_verify_ssl', None, os.environ.get('SITE_SYNC_VERIFY_SSL', '1'))
        if delivery_cost:
            form.delivery_cost.data = float(delivery_cost)
        if markup_percentage:
            form.markup_percentage.data = float(markup_percentage)
        form.site_sync_endpoint.data = site_sync_endpoint or ''
        form.site_sync_token.data = site_sync_token or ''
        form.site_sync_host_header.data = site_sync_host_header or ''
        form.site_sync_verify_ssl.data = str(site_sync_verify_ssl).lower() not in ('0', 'false', 'no', 'off')
    
    return render_template('settings/index.html', form=form)

# Export routes

@app.route('/export', methods=['GET', 'POST'])
def export_index():
    """Export page"""
    form = ExportForm()
    
    if form.validate_on_submit():
        try:
            # Build query based on selected category (only published bouquets)
            query = Bouquet.query.filter_by(published=True)
            if form.category_id.data:
                query = query.filter_by(category_id=form.category_id.data, published=True)
            
            bouquets = query.order_by(Bouquet.sku).all()
            
            # Create Excel workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "Прайс-лист букетов"
            
            # Define styles
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center")
            
            price_font = Font(bold=True)
            center_alignment = Alignment(horizontal="center")
            right_alignment = Alignment(horizontal="right")
            
            # Write header
            ws['A1'] = 'Артикул'
            ws['B1'] = 'Цена (лей)'
            
            # Style header
            for cell in ws[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
            
            # Set column widths
            ws.column_dimensions['A'].width = 15
            ws.column_dimensions['B'].width = 15
            
            # Write data
            for row, bouquet in enumerate(bouquets, start=2):
                # Round price to nearest 10
                rounded_price = round(bouquet.final_price / 10) * 10
                
                ws[f'A{row}'] = bouquet.sku
                ws[f'B{row}'] = int(rounded_price)
                
                # Style data cells
                ws[f'A{row}'].alignment = center_alignment
                ws[f'B{row}'].alignment = right_alignment
                ws[f'B{row}'].font = price_font
            
            # Save to BytesIO
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)
            
            # Create response
            response = make_response(output.getvalue())
            response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            
            # Generate filename: "Название категории - количество.xlsx"
            bouquet_count = len(bouquets)
            if form.category_id.data:
                category = Category.query.get(form.category_id.data)
                category_name = category.name if category else 'Неизвестная категория'
                filename = f'{category_name} - {bouquet_count}.xlsx'
            else:
                filename = f'Все букеты - {bouquet_count}.xlsx'
            
            # Properly encode filename for Content-Disposition with UTF-8 support
            encoded_filename = quote(filename)
            response.headers['Content-Disposition'] = f"attachment; filename*=UTF-8''{encoded_filename}"
            
            return response
            
        except Exception as e:
            flash('Error generating export. Please try again.', 'error')
            app.logger.error(f'Error generating export: {e}')
    
    return render_template('export/index.html', form=form)

# API endpoints for dynamic functionality



@app.route('/api/bouquet/<int:id>/price')
def api_bouquet_price(id):
    """Calculate bouquet price dynamically"""
    bouquet = Bouquet.query.get_or_404(id)
    return {
        'final_price': bouquet.calculate_price()
    }




@app.route('/api/recalculate-bouquet-prices', methods=['POST'])
def api_recalculate_bouquet_prices():
    """Manually recalculate all bouquet prices using SQL UPDATE"""
    try:
        from models import recalculate_bouquet_prices
        
        # Get total bouquets count before update
        total_bouquets = Bouquet.query.count()
        
        # Recalculate ALL bouquets using SQL UPDATE
        updated_count = recalculate_bouquet_prices()
        
        return jsonify({
            'success': True,
            'updated_count': updated_count,
            'total_bouquets': total_bouquets,
            'message': f'Успешно пересчитано {updated_count} букетов из {total_bouquets}'
        })
        
    except Exception as e:
        db.session.rollback()
        app.logger.error(f'Error recalculating bouquet prices: {e}')
        return jsonify({
            'success': False,
            'error': str(e),
            'message': 'Ошибка при пересчете цен. Попробуйте еще раз.'
        }), 500


@app.route('/api/site-sync-prices', methods=['POST'])
def api_site_sync_prices():
    """Preview or send CRM bouquet prices to Joomla/VirtueMart by SKU."""
    try:
        data = request.get_json(silent=True) or {}
        mode = data.get('mode', 'preview')
        category_id = data.get('category_id')
        search_query = (data.get('search') or '').strip()
        published_filter = data.get('published') or ''

        try:
            category_id = int(category_id) if category_id else None
        except (TypeError, ValueError):
            category_id = None

        query = build_bouquet_filter_query(category_id, search_query, published_filter)
        bouquets = query.order_by(Bouquet.sku).all()

        items = []
        missing_sku = []

        for bouquet in bouquets:
            new_price = bouquet.calculate_price()
            bouquet.final_price = new_price

            if not bouquet.sku:
                missing_sku.append({
                    'id': bouquet.id,
                    'name': bouquet.name
                })
                continue

            items.append({
                'sku': bouquet.sku.strip(),
                'name': bouquet.name,
                'price': int(round(new_price / 10) * 10),
                'category': bouquet.category.name if bouquet.category else '',
                'published': bool(bouquet.published)
            })

        db.session.commit()

        summary = {
            'total_matched': len(bouquets),
            'ready_to_sync': len(items),
            'missing_sku': missing_sku,
            'category_id': category_id,
            'search': search_query,
            'published': published_filter
        }

        if mode == 'preview':
            return jsonify({
                'success': True,
                'mode': 'preview',
                'summary': summary,
                'items': items[:50],
                'truncated': len(items) > 50,
                'configured': bool(get_site_sync_settings()['endpoint'] and get_site_sync_settings()['token'])
            })

        sync_result = verify_live_site_prices(items, sync_prices_to_site(items))

        return jsonify({
            'success': bool(sync_result.get('success')),
            'mode': 'sync',
            'summary': summary,
            'site': sync_result
        }), 200 if sync_result.get('success') else 400

    except Exception as e:
        db.session.rollback()
        app.logger.error(f'Error syncing prices to site: {e}')
        return jsonify({
            'success': False,
            'message': 'Ошибка синхронизации цен с сайтом.',
            'error': str(e)
        }), 500


@app.route('/api/export-stats')
def api_export_stats():
    """Get export statistics with pagination"""
    try:
        category_id = request.args.get('category_id', type=int)
        page = request.args.get('page', 1, type=int)
        per_page = 50  # Show 50 items per page
        
        # Build query based on category filter (only published bouquets)
        query = Bouquet.query.filter_by(published=True)
        if category_id:
            query = query.filter_by(category_id=category_id, published=True)
        
        # Get paginated results
        bouquets_paginated = query.order_by(Bouquet.sku).paginate(
            page=page, 
            per_page=per_page, 
            error_out=False
        )
        
        total_bouquets = bouquets_paginated.total
        
        # Calculate average price from all bouquets (not just current page)
        if total_bouquets > 0:
            all_bouquets = query.all()
            avg_price = sum(bouquet.final_price for bouquet in all_bouquets) / total_bouquets
            avg_price = round(avg_price / 10) * 10  # Round to nearest 10
        else:
            avg_price = 0
        
        # Get total categories
        total_categories = Category.query.count()
        
        return jsonify({
            'success': True,
            'total_bouquets': total_bouquets,
            'total_categories': total_categories,
            'avg_price': int(avg_price),
            'bouquets': [{'sku': b.sku, 'name': b.name, 'price': round(b.final_price / 10) * 10} for b in bouquets_paginated.items],
            'pagination': {
                'page': page,
                'pages': bouquets_paginated.pages,
                'per_page': per_page,
                'total': total_bouquets,
                'has_prev': bouquets_paginated.has_prev,
                'has_next': bouquets_paginated.has_next,
                'prev_num': bouquets_paginated.prev_num,
                'next_num': bouquets_paginated.next_num
            }
        })
        
    except Exception as e:
        app.logger.error(f'Error getting export stats: {e}')
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

# User management routes
@app.route('/user')
def user_index():
    """User management page"""
    user = User.query.get(session.get('user_id'))
    if not user:
        return redirect(url_for('login'))
    return render_template('user/index.html', user=user)

@app.route('/user/edit', methods=['GET', 'POST'])
def user_edit():
    """Edit user credentials"""
    form = UserForm()
    user = User.query.get(session.get('user_id'))
    if not user:
        return redirect(url_for('login'))
    
    # Pre-fill username
    if request.method == 'GET':
        form.username.data = user.username
    
    if form.validate_on_submit():
        # Check if username is already taken by another user
        existing_user = User.query.filter(User.username == form.username.data, User.id != user.id).first()
        if existing_user:
            flash('Этот логин уже используется', 'error')
            return render_template('user/edit.html', form=form)
        
        try:
            # Update user data
            user.username = form.username.data
            new_password = (form.new_password.data or '').strip()
            confirm_password = (form.confirm_password.data or '').strip()

            if new_password:
                if new_password != confirm_password:
                    flash('Пароли должны совпадать', 'error')
                    return render_template('user/edit.html', form=form)
                user.set_password(new_password)

            db.session.commit()
            
            session['username'] = user.username
            session['role'] = user.role
            flash('Данные пользователя успешно обновлены!', 'success')
            return redirect(url_for('user_index'))
            
        except Exception as e:
            db.session.rollback()
            flash('Ошибка при обновлении данных. Попробуйте снова.', 'error')
    
    return render_template('user/edit.html', form=form)

@app.route('/admin/users')

def admin_users():
    """Admin panel for user management"""
    # Отладочная информация
    app.logger.info(f"Текущий пользователь: {current_user.username}, роль: {current_user.role}")
    
    if current_user.role != 'admin':
        flash('Доступ запрещен. Только для администраторов.', 'error')
        return redirect(url_for('index'))
    
    # Получаем всех пользователей
    users = User.query.order_by(User.created_at.desc()).all()
    
    return render_template('admin/users.html', users=users)

@app.route('/admin/users/<int:user_id>/approve', methods=['POST'])

def admin_approve_user(user_id):
    """Approve a pending user"""
    if not current_user.is_admin():
        flash('Доступ запрещен. Только для администраторов.', 'error')
        return redirect(url_for('index'))
    
    try:
        user = User.query.get_or_404(user_id)
        
        if user.status != 'pending':
            flash('Пользователь уже обработан', 'warning')
            return redirect(url_for('admin_users'))
        
        # Одобряем пользователя
        user.status = 'active'
        user.approved_at = datetime.utcnow()
        user.approved_by = current_user.id
        
        db.session.commit()
        
        flash(f'Пользователь {user.username} успешно одобрен', 'success')
        
    except Exception as e:
        db.session.rollback()
        flash('Ошибка при одобрении пользователя', 'error')
        app.logger.error(f'Error approving user: {e}')
    
    return redirect(url_for('admin_users'))

@app.route('/admin/users/<int:user_id>/suspend', methods=['POST'])

def admin_suspend_user(user_id):
    """Suspend a user"""
    if not current_user.is_admin():
        flash('Доступ запрещен. Только для администраторов.', 'error')
        return redirect(url_for('index'))
    
    try:
        user = User.query.get_or_404(user_id)
        
        if user.is_admin():
            flash('Нельзя заблокировать администратора', 'error')
            return redirect(url_for('admin_users'))
        
        # Блокируем пользователя
        user.status = 'suspended'
        
        db.session.commit()
        
        flash(f'Пользователь {user.username} заблокирован', 'success')
        
    except Exception as e:
        db.session.rollback()
        flash('Ошибка при блокировке пользователя', 'error')
        app.logger.error(f'Error suspending user: {e}')
    
    return redirect(url_for('admin_users'))

@app.route('/admin/users/<int:user_id>/activate', methods=['POST'])

def admin_activate_user(user_id):
    """Activate a suspended user"""
    if not current_user.is_admin():
        flash('Доступ запрещен. Только для администраторов.', 'error')
        return redirect(url_for('index'))
    
    try:
        user = User.query.get_or_404(user_id)
        
        # Активируем пользователя
        user.status = 'active'
        if not user.approved_at:
            user.approved_at = datetime.utcnow()
            user.approved_by = current_user.id
        
        db.session.commit()
        
        flash(f'Пользователь {user.username} активирован', 'success')
        
    except Exception as e:
        db.session.rollback()
        flash('Ошибка при активации пользователя', 'error')
        app.logger.error(f'Error activating user: {e}')
    
    return redirect(url_for('admin_users'))
